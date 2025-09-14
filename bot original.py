import asyncio
import random
import re
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, ContextTypes,
    MessageHandler, CallbackQueryHandler, filters
)

# 🔹 Charger l'Excel et corriger les caractères
paragraphs = {}
wb = load_workbook(r"C:\m\kt++.xlsx")
ws = wb.active

# Sauter l'en-tête
for row in ws.iter_rows(min_row=2, values_only=True):
    numero, texte = row
    if numero is None or texte is None:
        continue

    numero = str(int(numero))
    texte = str(texte).strip()

    # 🔧 Correction des accents mal encodés
    try:
        texte = texte.encode("latin1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass

    paragraphs[numero] = texte

print(f"{len(paragraphs)} paragraphes chargés.")

# 🔹 Fonction pour envoyer un texte long
async def send_long_text(chat, text):
    max_len = 4000
    for i in range(0, len(text), max_len):
        await chat.send_message(text[i:i+max_len], parse_mode="HTML")

# 🔹 /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message:
        await update.message.reply_text(
            "Bonjour ! Envoyez un numéro de paragraphe ou utilisez les commandes :\n"
            "/help - liste des commandes\n"
            "/search <mot> - recherche un mot\n"
            "/range <début>-<fin> - affiche une plage de paragraphes\n"
            "/random - paragraphe aléatoire\n"
            "/stats [mot] - statistiques\n"
            "/credits - infos sur le bot"
        )

# 🔹 /help
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message:
        await update.message.reply_text(
            "Commandes disponibles :\n"
            "1. Envoyer un numéro → obtient le paragraphe correspondant\n"
            "2. /help → affiche ce message\n"
            "3. /search <mot(s)> → cherche un ou plusieurs mots dans les paragraphes\n"
            "4. /range <début>-<fin> → affiche tous les paragraphes dans cette plage\n"
            "5. /random → renvoie un paragraphe aléatoire\n"
            "6. /stats [mot] → statistiques globales ou sur un mot\n"
            "7. /credits → infos sur le bot"
        )

# 🔹 /search
async def search_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2:
        await update.message.reply_text("Usage : /search <mot(s)>")
        return
    keywords = parts[1].lower().split()
    results = []
    for num, txt in paragraphs.items():
        lower_txt = txt.lower()
        if all(k in lower_txt for k in keywords):
            highlighted = txt
            for k in keywords:
                highlighted = re.sub(
                    f"({re.escape(k)})", r"<b>\1</b>", highlighted, flags=re.IGNORECASE
                )
            results.append(f"Paragraphe {num} : {highlighted}")
    if results:
        for res in results[:10]:
            await send_long_text(update.message.chat, res)
        if len(results) > 10:
            await update.message.reply_text(f"...et {len(results)-10} autres résultats.")
    else:
        await update.message.reply_text("Aucun paragraphe trouvé avec ces mots.")

# 🔹 /range
async def range_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2 or "-" not in parts[1]:
        await update.message.reply_text("Usage : /range <début>-<fin>")
        return
    try:
        start_num, end_num = map(int, parts[1].split("-"))
    except ValueError:
        await update.message.reply_text("Les nombres doivent être valides.")
        return
    if start_num > end_num:
        start_num, end_num = end_num, start_num
    for n in range(start_num, end_num + 1):
        key = str(n)
        if key in paragraphs:
            await send_long_text(update.message.chat, f"Paragraphe {key} :\n{paragraphs[key]}")
        else:
            await update.message.reply_text(f"Paragraphe {n} non trouvé.")

# 🔹 /random
async def random_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return
    key = random.choice(list(paragraphs.keys()))
    keyboard = [
        [
            InlineKeyboardButton("⬅️ Précédent", callback_data=f"{int(key)-1}"),
            InlineKeyboardButton("➡️ Suivant", callback_data=f"{int(key)+1}")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await send_long_text(update.message.chat, f"Paragraphe {key} :\n{paragraphs[key]}")
    await update.message.reply_text("Navigation :", reply_markup=reply_markup)

# 🔹 /stats
async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return
    parts = update.message.text.strip().split(maxsplit=1)
    if len(parts) == 1:
        total = len(paragraphs)
        avg_words = sum(len(txt.split()) for txt in paragraphs.values()) / total
        await update.message.reply_text(
            f"📊 Statistiques globales :\n"
            f"- Nombre total de paragraphes : {total}\n"
            f"- Nombre moyen de mots par paragraphe : {avg_words:.1f}"
        )
    else:
        keyword = parts[1].lower()
        para_count = sum(1 for txt in paragraphs.values() if keyword in txt.lower())
        word_count = sum(txt.lower().count(keyword) for txt in paragraphs.values())
        await update.message.reply_text(
            f"📊 Statistiques pour le mot '<b>{keyword}</b>' :\n"
            f"- Apparaît dans {para_count} paragraphes\n"
            f"- Nombre total d'occurrences : {word_count}",
            parse_mode="HTML"
        )

# 🔹 /credits
async def credits_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message:
        await update.message.reply_text("🤖 Ce bot a été créé en 2025 par @LNJ21.")

# 🔹 Navigation boutons
async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    try:
        await query.answer(cache_time=1)  # éviter l’erreur “Query too old”
    except:
        pass

    key = str(int(query.data))
    if key in paragraphs:
        keyboard = [
            [
                InlineKeyboardButton("⬅️ Précédent", callback_data=f"{int(key)-1}"),
                InlineKeyboardButton("➡️ Suivant", callback_data=f"{int(key)+1}")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await send_long_text(query.message.chat, f"Paragraphe {key} :\n{paragraphs[key]}")
        await query.message.reply_text("Navigation :", reply_markup=reply_markup)
    else:
        await query.message.reply_text("Paragraphe non trouvé.")

# 🔹 Message avec numéro direct
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return
    text = update.message.text.strip()
    if text.isdigit():
        key = str(int(text))
        if key in paragraphs:
            keyboard = [
                [
                    InlineKeyboardButton("⬅️ Précédent", callback_data=f"{int(key)-1}"),
                    InlineKeyboardButton("➡️ Suivant", callback_data=f"{int(key)+1}")
                ]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await send_long_text(update.message.chat, f"Paragraphe {key} :\n{paragraphs[key]}")
            await update.message.reply_text("Navigation :", reply_markup=reply_markup)
        else:
            await update.message.reply_text("Numéro de paragraphe inconnu.")
    else:
        await update.message.reply_text("Merci d'envoyer un numéro ou une commande valide (/help).")

# 🔹 Lancer le bot
if __name__ == "__main__":
    TOKEN = "8303903539:AAF9uP0x9ntBfkG7V26WGEiYmQxjYX5DwDo"
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("search", search_command))
    app.add_handler(CommandHandler("range", range_command))
    app.add_handler(CommandHandler("random", random_command))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CommandHandler("credits", credits_command))
    app.add_handler(CallbackQueryHandler(button_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    print("Bot démarré...")
    asyncio.run(app.run_polling())
